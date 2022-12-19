#note043_excel_cell2.py 
#prepare libraries and settings

from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active


########## 1. APPEND VALUES TO CELL (PER LINE)
# append values in list format
ws.append(["id", "eng", "math"])
for i in range(1,11):
  ws.append([i, randint(0, 100), randint(0,100)])
ws.save("sample.xlsx")


########## 2. CHECK VALUES
#1) check particular column's data
# let's assume you are an ENG teacher. You will only need to check column B's data
col_b = ws["B"]
print(col_b)        #result: (<Cell 'Sheet'.B1>, <Cell 'Sheet'.B2>... <Cell 'Sheet'.B11>

for cell in col_b:
  print(cell.value) #result: column's values

#2) check multiple column's data
col_range = ws["B":"C"]
for cols in col_range:    #start from 2nd column and then go to 3rd column
  for cell in cols:
    print(cell.value)     #result: eng columns and then math column
    
#3) check row's data
row_title = ws[1]
for cell in row_title:
  print(cell.value)       #result: id, eng, math

#4) check mulitple row's data
row_range = ws[2:6]       #from 2nd row to 6th row (=! slicing, incl 6)
for rows in row_range:
  for cell in rows:
    print(cell.value, end =" ") 
 
#5) in reality, we may not know how many rows are there
row_range = ws[2:max_row]
for rows in row_range:
  for cell in rows:
    print(cell.value, end =" ")
    
    
########## 3. GET CELL'S INFO TOGETHER WITH VALUE
#1) check values with coordinate cell info
row_range = ws[2:max_row]

from openpyxl.utils.cell import coordinate_from_string
for rows in row_range:
  for cell in rows:
    print(cell.value, end = " ")        #this will give cell's value
    print(cell.coordinate, end = " ")   #this will give cell's info (A2, B2, C2...)
  print()
  
#result:
#  1 A2 22 B2 34 C2 
#  2 A3 99 B3 25 C3...

#2) check values with coordiante cell info in tuple form (easier to read)
from openpyxl.utils.cell import coordinate_from_string
for rows in row_range:
  for cell in rows:
    xy = coordinate_from_string(cell.coordinate)
    print(xy, end= " ")
  print()
  
#result:
#  ('A', 2) ('B', 2) ('C', 2) 
#  ('A', 3) ('B', 3) ('C', 3)...

#3) check values with coordiante cell info in tuple form (separately for rows/columns)
from openpyxl.utils.cell import coordinate_from_string
for rows in row_range:
  for cell in rows:
    xy = coordinate_from_string(cell.coordinate)
    print(xy[0], end="")    #A as in from A1
    print(xy[1], end= " ")  #1 as in from A1
    
#result:
#   A2 B2 C2 
#   A3 B3 C3 
    

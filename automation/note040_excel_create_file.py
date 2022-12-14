#note040_excel_create_file.py


########## 1. RPA (Robotic Process Automation)
#1) Scenario: You need to print name card for an event wher 1,000 people will join
#             For the name card, you will need to copy/paste attendee's name, company name and phone number from the list in Excel
#             Think about the task. You copy and paste manually and repetitively for 3 parts, 1,000 times.
#             What if we can automate this process?
#2) Benefit:  Save time, reduce human error


########## 2. OPENPYXL LIBRARY
# install openpyxl library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files
# ref: https://openpyxl.readthedocs.io/en/latest/tutorial.html
# write below on the terminal:
pip install openpyxl


########## 3. EXCEL: CREATE FILE
#1) create new workbook, use active sheet, add title
from openpyxl import Workbook
wb = Workbook()
wb = wb.active            #bring the activated sheet
wb.title = "sheet_1213"   #change sheet's name

#2) need to archive so that the file is created
wb.save("sameple.xlsx")   #save file under the title 'sample'
wb.close()

##what happens? nothing happens in the terminal but the file will be created on the VSC's EXPLORER section.


########## 4. EXCEL: CREATE SHEET
#1) create new workbook, create new sheet, add title
from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()    #create new sheet (additionally)
ws.title = "my_sheet"     #assign sheet's name

#2) or directly assign sheet's name at the time of creation
ws1 = wb.create_sheet("your_sheet")

#3) if you want to create sheet at specific location, use index number
ws1 = wb.create_sheet("new_sheet", 2)    #sheet will be create as 3rd tab (index number starts from 0, 1, 2...) 

#4) select sheet
new_ws = wb["new_sheet"]

#5) get sheet's name
print(wb.sheetnames)
# or
for sheet in wb:
  print(sheet.title)

#6) add tab's color (sheet's tab)
ws.sheet_properties.tabColor = "ff66ff"     #check the color from here: https://www.w3schools.com/colors/colors_rgb.asp


########## 5. EXCEL: DUPLICATE SHEET
#) copy and paste (duplicate) the sheet
new_ws["A1"] = "Sample Sentence"    #on the cell A1 from sheet "new_sheet", write "Sample Sentence" 
target = wb.copy_worksheet(new_ws)  #copy "new_sheet" -> save into variable 
target.title = "dup_sheet"          #assign title to the copied sheet as "dup_sheet"
wb.save("sample.xlsx")              #save file

  

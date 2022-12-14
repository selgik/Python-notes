#note041_excel_cell.py
#prepare libraries and settings

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "TestSheet"
wb.save("sample.xlsx")    #in order to check the result in Excel file, always save.


########## 1. INPUT VALUES TO CELL
#1) by accessing to cell directly
ws["A1"] = 1      #input 1 to A1
ws["A2"] = 2      #input 2 to A2
ws["A3"] = 3      #input 3 to A3

#2) by accessing column/row number
ws.cell(column=3, row=1, value=10)    #input 10 to C1


########## 2. READ VALUES FROM CELL
#1) difference between printing cell and cell value
print(ws["A1"])         #result: <Cell 'TestSheet' .A1> ---> gives object information about cell
print(ws["A1"].value)   #result: 1 --> gives value in A1
print(ws["A10"].value)  #result: None 

#2) read with column/row information
print(ws.cell(column=, row=2).value)  


########## 3. FILL IN CELLS WITH LOOP
#1) how does loop works:
index = 0
for x in range(1, 11):
  for y in range(1, 11):
    ws.cell(row=x, column=y, value=index)
    index +=1
##result:
##A1:0  B1:1  C1:2  D1:3   E1:4   F1:5   G1:6   H1:7   I1:8   J1:9  
##A2:10 B2:11 C1:12 D2:13  E2:14  F2:15  G2:16  H2:17  I2:18  J2:19  
##...

#2) fill in with random number:
from random import *
for x in range(1,11):
  for y in range(1,11):
    ws.cell(row=x, column=y, value=randint(1,100))  #from A1 to J10, fill in with random number between 1-100

wb.save("sample.xlsx")

  

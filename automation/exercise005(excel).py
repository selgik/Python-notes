#exercise005(excel).py
#prepare libraries and settings

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

##### Scenario:
# You are an assistant working for a professor X. 
# Professor X needs to prepare excel file with all grade details.
# Using below data, complete tasks to help professor X.


##### 1. MY VERSION
#1) Add data to the excel
data = (
  ("ID", "Attnd", "Q1", "Q2", "MidT", "FinT", "Prj"),
  (1,10,8,5,14,26,12),
  (2,7,3,7,15,24,18),
  (3,9,5,8,8,12,4),
  (4,7,8,7,17,21,18),
  (5,7,8,7,16,25,15),
  (6,3,5,8,8,17,0),
  (7,4,9,10,16,27,18),
  (8,6,6,6,15,19,17),
  (9,10,10,9,19,30,19),
  (10,9,8,8,20,25,20)
)

for d in data:
  ws.append(d)

#2) Quiz2 contained error. We need to update all students' scores to 10 under column D
for col in ws.iter_cols(min_col=4, max_col=4):
  for i in range(1, 11):
    col[i].value = 10

#3) Using SUM formula, obtain total scores
for sm in ws.iter_cols(min_col=8, max_col=8):
  for j in range(1,12):
    sm[j-1].value = "=SUM(B"+str(j)+":G"+str(j)+")"
    
ws["H1"] = "Total"

#4) Based on the criteria, give scores
#a. If total scores > 90 then A
#b. If total scores > 80 then B
#c. If total scores > 70 then C
#d. Otherwise D
#e. Regardless of scores, F if column B (Attendance) score is less than 5

#first give scores (before running, need to save file so that column H value can be read)
for v in range(2,12):
  if ws["H"+str(v)].value >= 90:
    ws["I"+str(v)] = "A"
  if ws["H"+str(v)].value >= 80 and ws["H"+str(v)].value < 90:
    ws["I"+str(v)] = "B"
  if ws["H"+str(v)].value >= 70 and ws["H"+str(v)].value < 80:
    ws["I"+str(v)] = "C"
  else:
    ws["I"+str(v)] = "D"

#override with F if attendance score is less than 5
for v in range(2,12):
  if ws["B"+str(v)].value > 5:
    continue
  else:
    ws["I"+str(v)] = "F"

ws["I1"] = "Grade"


##### 2. CODE REVIEW
# My codes worked out, but they are hard to read and messy especially #3 comparing to instructor's codes.
# Not a perfect solution for #4 as I had to open, save, close file so that value is saved before running the code.


##### 3. INSTRUCTOR'S VERSION
#1) Add data to the excel
ws.append(("ID", "Attnd", "Q1", "Q2", "MidT", "FinT", "Prj"))

scores = [
  (1,10,8,5,14,26,12),
  (2,7,3,7,15,24,18),
  (3,9,5,8,8,12,4),
  (4,7,8,7,17,21,18),
  (5,7,8,7,16,25,15),
  (6,3,5,8,8,17,0),
  (7,4,9,10,16,27,18),
  (8,6,6,6,15,19,17),
  (9,10,10,9,19,30,19),
  (10,9,8,8,20,25,20)
]

for s in scores:
  ws.append(s)

#Q: why instructor appended title and then saved scores in the list format?
#A: later, instructor uses variable 'scores' in for loop.

#2) Quiz2 contained error. We need to update all students' scores to 10 under column D
for idx, cell in enumerate(ws["D"]):
    if idx == 0:                      #idx = 0 is title, which is "Q2"
       continue                       #so skip if idx = 0
    cell.value = 10                   #only override value with 10 for the rest

#3) Using SUM formula, obtain total scores
for idx, score in enumerate(scores, start = 2):                           #why start=2? in Excel, 1st row is title so data will be inserted from 2nd row
  ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})" .format(idx, idx)    #row=idx which starts from 2.
  sum_val = sum(score[1:]) - score[3] + 10                                #score[0] is ID and should be excluded for total SUM calculation
                                                                          #this variable will be used later
  
#4) Based on the criteria, give scores
#a. If total scores > 90 then A
#b. If total scores > 80 then B
#c. If total scores > 70 then C
#d. Otherwise D
#e. Regardless of scores, F if column B (Attendance) score is less than 5

grade = None
if sum_val >= 90:
  grade = "A"
if sum_val >= 80:
  grade = "B"
if sum_val >= 70:
  grade = "C"
else:
  grade = "D"
  
if score[1].value < 5:
  grade = "F"

ws.cell(row=idx, column=9).value = grade

#because sum_val was created, there is no need to open, save, close Excel file so that formula gets evaluated. 


wb.save("scores.xlsx")

  

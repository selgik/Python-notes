#note045_excel_modify.py 
#prepare libraries and settings

from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


########## 1. SEARCH
#conditional search
for row in ws.iter_rows(min_row=2):                   #by limiting data as min_row=2, exclude title row
  if int(row[1].value) > 80:                          #if eng score is above 80
    print(row[0].value, "th id is above 80 score")    #print (student number) is above 80 score

    
########## 2. OVERRIDE
#override data(title)
for row in iter_rows(max_row=1):                      #by limiting data as title row
  for cell in row:                                    #review cells 
    if cell.value == "eng":                           #if cell value is "eng"
      cell.value = "computer"                         #update to "computer"
  

########## 3. INSERT
#1) insert rows
ws.insert_rows(8)                                     #empty 8th row will be created
ws.insert_rows(8, 5)                                  #empty 5lines will be inserted from the 8th row

#2) insert columns
ws.insert_cols(2)                                     #empty column will be inserted in the 2nd (B) column
ws.insert_cols(2, 3)                                  #empty 3lines will be inserted from the 2nd (B) column
ws.save("sample_insert.xlsx")


########## 4. DELETE
#1) delete row
ws.delete_rows(8)                                     #8th row (id=7 student) data will be delted, (after id=6, id=8 will come)
ws.delete_rows(8, 3)                                  #3lines from 8th row will be deleted

#2) delete column
ws.delete_cols(2)                                     #2nd (B) column will be deleted (after id column, math column will come)
ws.delete_cols(2, 2)                                  #2lines from 2nd (B) column will be deleted
ws.save("sample_delete.xlsx")
  

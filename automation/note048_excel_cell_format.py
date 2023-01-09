#note048_excel_cell_format.py
#prepare libraries and settings

from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


########## 1. PREPARE
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


########## 2. FORMAT CHANGE
#1) define variable
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

#2) adjust row's height or column's width
ws.row_dimensions[1].height = 50
ws.column_dimensions["A"].width = 5

#3) change font property
a1.font = Font(color="FF0000", italic=True, bold=True)        #use rbg color
b1.font = Font(color="CC33FF", name="Arial", strike=True)     #apply strike line
c1.font = Font(color="0000FF", size=20, underline = "single") #apply underline

#4) border lines
##result: A1, B1, C1 cell's border will be applied with thin line
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

#5) align all cells to the center
for row in ws.rows:   
  for cell in row:    
    cell.alignemnt = Alignment(horizontal = "center", vertial ="center")  #other options: right, left, top, bottom

#6) freeze row or column
ws.freeze_panes = "B2" 

##result: freeze row1 and column1 
##        B2 will become a point where row/column freeze will be applied. 
##        Based on B2, left column (which is column A) and above row (which is 1st row) will be frozen.

#7) merge/unmerge cells
ws.merge_cells("E2:G2")
ws["E2"].value = "Merged cells"
ws.unmerge_cells("E2:G2")


########## 3. CONDITIONAL FORMATIING
# task: for those cells where its value is >= 90, hightlight cell with green and change font to red color

for row in ws.rows:       #let's bring all rows in tuple format
  for cell in row:        #within those tuples, let's bring cell one by one
    if cell.column == 1:  #1st column which is ID column is not needed here so 'continue'
      continue            
    if isinstance(cell.value, int) and cell.value >= 90:                #if type is int (exclude string) and its value >=90
      cell.fill = PatternFill(fgColor="00FF00", fill_type = "solid")    #then fill in with green color
      cell.font = Font(color="FF0000")                                  #change font color to red

ws.save("sample_styles.xlsx")

  

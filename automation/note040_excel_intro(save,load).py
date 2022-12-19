#note040_excel_intro.py


########## 1. RPA (Robotic Process Automation)
#1) Scenario: You need to print name card for an event wher 1,000 people will join
#             For the name card, you will need to copy/paste attendee's name, company name and phone number from the list in Excel
#             Think about the task. You copy and paste manually and repetitively for 3 parts, 1,000 times.
#             What if we can automate this process?
#2) Benefit:  Save time, reduce human error


########## 2. OPENPYXL LIBRARY
# install openpyxl library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files
# ref: https://openpyxl.readthedocs.io/en/latest/tutorial.html
# write below on the terminal:
pip install openpyxl


########## 3. EXCEL: CREATE FILE
#1) create new workbook, use active sheet, add title
from openpyxl import Workbook
wb = Workbook()
wb = wb.active            #bring the activated sheet
wb.title = "sheet_1213"   #change sheet's name

#2) need to archive so that the file is created
wb.save("sameple.xlsx")   #save file under the title 'sample'
wb.close()

##what happens? nothing happens in the terminal but the file will be created on the VSC's EXPLORER section.


########## 4. EXCEL: LOAD FILE
# assuming that data has been added through note042_excel_cell.py, let's load data
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active  #activated sheet

#1) load data, knowing that we have data within table 10 x 10
for x in range(1,11):
  for y in range(1,11):
    print(ws.cell(row=x, column=y).value, end=" ")
  print()
  
#2) load data, assuming that we do not know how many rows/columns are there to pull out
for x in range(1, ws.max_row+1):
  for y in range(1, ws.max_column+1):
    print(ws.cell(row=x, column=y).value, end= " ")
  print()

  

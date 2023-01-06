#note047_excel_formula.py
#prepare libraries and settings

import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active


########## 1. USE FORMULA
ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1,2,3)"      #6
ws["A3"] = "=AVERAGE(1,2,3)"  #2
ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"      #30

ws.save("sample_formula.xlsx")


########## 2. READ DATA
from openpyxl import load_workbook
wb = load_workbook("sample_formula.xlsx")
ws = wb.active

for row in ws.values:   
  for cell in row:    
    print(cell)       
    
#result: formulas will appear
#2023-01-07 00:00:00
#SUM(1,2,3)
#AVERAGE(1,2,3)
#10
#20
#=SUM(A4:A5)


########## 3. READ DATA(no formula but numbers only)
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

for row in ws.values:
  for cell in row:
    print(cell)
   
#result: formulas will appear
#2023-01-07 00:00:00
#None
#None
#10
#20
#None

### Why None?
#when we create a file via openpyxl, formulas are not processed. Formulas are just inserted into the cells.
#When Excel file gets opened, formulas will then be processed and results will show up.
#so 'None' will appear because there is no data processed (it's not evaluated). This is not a bug.

### So how to tackle it?
#open file manually, save it. Then formulas will be processed and saved as numbers. Re-run the code.

  

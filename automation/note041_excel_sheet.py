#note041_excel_sheet.py
#prepare libraries and settings

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "TestSheet"
wb.save("sample.xlsx")    #in order to check the result in Excel file, always save.


########## 1. CREATE SHEET
#1) create new workbook, create new sheet, add title
from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()    #create new sheet (additionally)
ws.title = "my_sheet"     #assign sheet's name

#2) or directly assign sheet's name at the time of creation
ws1 = wb.create_sheet("your_sheet")

#3) if you want to create sheet at specific location, use index number
ws1 = wb.create_sheet("new_sheet", 2)    #sheet will be create as 3rd tab (index number starts from 0, 1, 2...) 

#4) select sheet
new_ws = wb["new_sheet"]

#5) get sheet's name
print(wb.sheetnames)
# or
for sheet in wb:
  print(sheet.title)

#6) add tab's color (sheet's tab)
ws.sheet_properties.tabColor = "ff66ff"     #check the color from here: https://www.w3schools.com/colors/colors_rgb.asp


########## 2. DUPLICATE SHEET
#) copy and paste (duplicate) the sheet
new_ws["A1"] = "Sample Sentence"    #on the cell A1 from sheet "new_sheet", write "Sample Sentence" 
target = wb.copy_worksheet(new_ws)  #copy "new_sheet" -> save into variable 
target.title = "dup_sheet"          #assign title to the copied sheet as "dup_sheet"
wb.save("sample.xlsx")              #save file

  

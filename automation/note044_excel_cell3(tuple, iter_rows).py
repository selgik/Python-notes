#note044_excel_cell3(iter_rows).py
#prepare libraries and settings

from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


########## 1. CHECK VALUES (PER ROW, PER COLUMN WITH TUPLE)
#1) ws.rows
print(ws.rows)                    #result: <generator object Worksheet._cells_by_row at 0x0000021560F7B610>

#2) tuple(ws.rows)
print(tuple(ws.rows))             #result: ((<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>,<Cell 'Sheet'.C1>), (<Cell 'Sheet'.A2> ...

for row in tuple(ws.rows):        #result: (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>,<Cell 'Sheet'.C1>), (<Cell 'Sheet'.A2> ...
  print(row)

for row in tuple(ws.rows):
  print(row[1].value)             #result: column B's value (eng scores) will be listed
                                  # why? : 1st index from tuple(ws.rows) is <Cell 'Sheet.B1> which is eng column (column B). 
                                  #        With loop, you get rows' data for eng column (<Cell 'Sheet.B1>, <Cell 'Sheet.B2> ... <Cell 'Sheet.B11>)

#3) tuple(ws.columns)   
print(tuple(ws.columns))          #result: same as above, but tuple will group data per column

for column in tuple(ws.columns):  
  print(column.value)             #result: (<Cell 'Sheet'.A1>, <Cell 'Sheet'.A2>,<Cell 'Sheet'.A3> ... <Cell 'Sheet'.A11>), (<Cell 'Sheet'.B1>...+

for column in tuple(ws.columns):
  print(column[0].value)          #result: id eng math
                                  # why? : 0th index from tuple(ws.columns) is <Cell 'Sheet'.A1> which is title row (id, eng, math)
                                  #        With loop, you get column's data for title row (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>)
  

########## 2. ITER_ROWS, ITER_COLS
#1) iter_rows
for row in ws.iter_rows():
  print(row)                      #result: same as 1-(2)

for row in ws.iter_rows():
  print(row[1].value)             #result: same as 1-(2), column B's value (eng scores) will be listed
    
#2) iter_cols
for col in ws.iter_cols():
  print(col)                      #result: same as 1-(3) 
  
for col in ws.iter_cols():
  print(col[0].value)             #result: same as 1-(3), id eng math
  
#3) benefit of iter_rows() and iter_cols()
# iter_rows(), iter_cols() allows you to limit data 
# tip1: if no min/max is defined, values will be auto-selected from data's min/max cell value
#   ex: for col in ws.iter_cols(min_row=2) <--- till max row and max column 
# tip2: what's the difference between iter_rows and iter_cols? iter_rows pulls data horizontally, iter_cols vertically.


#3-1) limit row range
for row in ws.iter_rows(min_row=1, max_row=5):
  print(row[2].value)              #result: show data from C1(min 1st row of column C) to C5(max 5th row of column C) !! min/max_row starts from 1

#3-2) limit row and column range (using iter_rows)
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
  print(row[0].value, row[1].value)#result: limit data from B2 to C11 and then show data where row index is 0 and 1--> B2 to C11
  
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
  print(row)                       #result: gives cell info (<Cell 'Sheet'.B2>, <Cell 'Sheet'.C2>) (<Cell 'Sheet'.B3>...)
  
#3-3) limit row and column range (using iter_cols)
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
  print(col)                       #result: gives cell info, from A1 to C5 (<Cell 'Sheet'.A1>, <Cell 'Sheet'.A2>)...(<Cell 'Sheet'.A5>) (<Cell 'Sheet'.B1>...)

  
ws.save("sample.xlsx")
  

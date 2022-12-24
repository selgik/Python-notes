#note045_excel_modify.py 
#prepare libraries and settings

from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


########## 1. SEARCH
#conditional search
for row in ws.iter_rows(min_row=2):                   #by limiting data as min_row=2, exclude title row
  if int(row[1].value) > 80:                          #if eng score is above 80
    print(row[0].value, "th id is above 80 score")    #print (student number) is above 80 score

    
########## 2. OVERRIDE
#override data(title)
for row in iter_rows(max_row=1):                      #by limiting data as title row
  for cell in row:                                    #review cells 
    if cell.value == "eng":                           #if cell value is "eng"
      cell.value = "computer"                         #update to "computer"
  

########## 3. INSERT


########## 4. DELETE

